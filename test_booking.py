import time
from datetime import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

# Constants
TEMPLATE_FILE = "movie_booking_test_template.xlsx"
RESULT_FILE = "movie_booking_test_results.xlsx"
URL = "http://127.0.0.1:5000"

# Set up Selenium WebDriver
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
driver.get(URL)
driver.maximize_window()
time.sleep(1)

# Load the Excel template
wb = load_workbook(TEMPLATE_FILE)
ws = wb.active

# Loop through each test case (starting from row 2)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    movie = row[2].value
    tickets = row[3].value
    expected_price = row[4].value

    try:
        # Refresh the page before each test
        driver.get(URL)
        time.sleep(1)

        # Select movie
        select_movie = Select(driver.find_element(By.NAME, "movie"))
        select_movie.select_by_visible_text(movie)

        # Enter ticket count
        ticket_input = driver.find_element(By.NAME, "tickets")
        ticket_input.clear()
        ticket_input.send_keys(str(tickets))

        # Submit the form
        driver.find_element(By.XPATH, "//button[text()='Book Tickets']").click()
        time.sleep(1)

        # Read the booking summary
        summary = driver.find_element(By.CLASS_NAME, "summary").text
        status = "PASS" if f"₹{expected_price}" in summary else "FAIL"

        # Write results back into Excel
        row[5].value = summary
        row[6].value = status
        row[7].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    except Exception as e:
        row[5].value = f"Error: {str(e)}"
        row[6].value = "FAIL"
        row[7].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Save the result Excel file
wb.save(RESULT_FILE)
driver.quit()
print(f"✅ Test completed. Results saved to '{RESULT_FILE}'.")
